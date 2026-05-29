"""local_xlsx provider adapter.

Wraps the existing openpyxl-based parser and apply logic behind the
ProviderAdapter interface.  Behaviour is unchanged — this is purely
a structural refactoring.
"""
from __future__ import annotations

import io
import tempfile
from pathlib import Path
from typing import Any

from openpyxl import load_workbook as _load
from openpyxl.comments import Comment

from ...domain import Operation
from .. import ProviderAdapter, ProviderCapabilities
from ...parser import parse_xlsx as _parse_xlsx


class LocalXlsxAdapter(ProviderAdapter):
    @property
    def provider_id(self) -> str:
        return "local_xlsx"

    @property
    def capabilities(self) -> ProviderCapabilities:
        return ProviderCapabilities(
            resource_kinds=["spreadsheet"],
            supports_read=True,
            supports_write=True,
            supports_comments=True,
            supports_versioning=True,
            supports_conflict_detection=True,
            supports_batch_apply=True,
            online=False,
        )

    def parse(self, raw: bytes, name: str = "workbook") -> dict[str, Any]:
        """Parse .xlsx bytes into a Workbook-compatible dict."""
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(raw)
            tmp_path = Path(tmp.name)
        try:
            wb = _parse_xlsx(tmp_path, name=name)
            return wb.model_dump()
        finally:
            tmp_path.unlink(missing_ok=True)

    def apply_operations(
        self,
        operations: list[Operation],
        base_raw: bytes,
        metadata: dict[str, Any] | None = None,
    ) -> bytes:
        """Apply operations to the base .xlsx bytes and return new bytes."""
        book = _load(io.BytesIO(base_raw), data_only=False, read_only=False)

        for op in operations:
            self._write_operation(book, op)

        out = io.BytesIO()
        book.save(out)
        return out.getvalue()

    def _write_operation(self, book, operation: Operation) -> None:
        """Write a single operation into the openpyxl workbook."""
        target = operation.target
        if not target.cell:
            raise ValueError(f"operation {operation.id} has no target cell")

        sheet_name = target.sheet
        if sheet_name and sheet_name not in book.sheetnames:
            raise ValueError(f"sheet {sheet_name!r} not found")
        ws = book[sheet_name] if sheet_name else book.active
        cell = ws[target.cell]

        if operation.kind == "add_cell_comment":
            cell.comment = Comment(
                operation.after.get("comment") or operation.rationale,
                "spreadbreadai",
            )
        elif operation.kind == "clear_cell":
            cell.value = None
        elif operation.kind in ("set_cell_value", "set_cell_formula"):
            cell.value = self._typed_operation_value(operation)
        else:
            raise ValueError(
                f"unsupported operation kind {operation.kind!r}"
            )

    def _typed_operation_value(self, operation: Operation) -> Any:
        if operation.kind == "set_cell_formula":
            return operation.after.get("formula")
        if operation.kind == "set_cell_value":
            return operation.after.get("value")
        raise ValueError(f"cannot convert operation {operation.kind!r} to cell value")
