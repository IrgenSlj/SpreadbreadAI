from __future__ import annotations

import re
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook as _load

from .domain import (
    Workbook,
    WorkbookNamedRange,
    WorkbookRisk,
    WorkbookSheet,
    new_workbook,
)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

_DEP_CAP = 500
_LIST_CAP = 50

# Pattern for sheet references like   Sheet1!A1   or   'Some Sheet'!A1
# Used as a fallback when the tokenizer is unavailable.
_SHEET_REF_RE = re.compile(r"'?([A-Za-z0-9_ ]+)'?!")

# Pattern for external workbook references: [...] inside a formula.
_EXTERNAL_RE = re.compile(r"\[([^\]]+)\]")

# Stale-input markers (case-insensitive, after strip)
_STALE_VALUES = {"todo", "tbd", "fixme", "?", "???", "xxx"}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _extract_range_tokens(formula: str) -> list[str]:
    """Return OPERAND/RANGE token values from *formula* using the openpyxl
    tokenizer.  Falls back to an empty list on any error."""
    try:
        from openpyxl.formula.tokenizer import Tokenizer  # type: ignore[import]

        return [
            tok.value
            for tok in Tokenizer(formula).items
            if tok.type == "OPERAND" and tok.subtype == "RANGE"
        ]
    except Exception:
        return []


def _sheet_names_from_tokens(tokens: list[str]) -> list[str]:
    """Extract sheet names referenced in a list of range tokens."""
    names: list[str] = []
    for tok in tokens:
        # Tokens like  Sheet1!A1  or  'Some Sheet'!A1  or  [File.xlsx]Sheet!A1
        if "!" in tok:
            prefix = tok.split("!")[0]
            # Strip external workbook part [File.xlsx]
            prefix = re.sub(r"\[[^\]]+\]", "", prefix)
            sheet = prefix.strip("'")
            if sheet:
                names.append(sheet)
    return names


def _external_file_from_formula(formula: str) -> Optional[str]:
    """Return the first external filename found in *formula*, or None."""
    m = _EXTERNAL_RE.search(formula)
    return m.group(1) if m else None


def _col_letter(n: int) -> str:
    """Convert a 1-based column number to a letter (up to ZZ)."""
    result = ""
    while n:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------


def parse_xlsx(path: Path, name: Optional[str] = None, owner: str = "user") -> Workbook:
    wb = new_workbook(name or path.stem, owner=owner)
    book = _load(path, data_only=False, read_only=False)  # read_only=False for defined_names

    known_sheets: set[str] = set(book.sheetnames)
    sheets: list[WorkbookSheet] = []
    risks: list[WorkbookRisk] = []
    dependencies: dict[str, list[str]] = {}

    for ws in book.worksheets:
        rows = ws.max_row or 0
        cols = ws.max_column or 0
        formula_cells = 0
        populated = 0
        sample: list[list[str]] = []

        ext_refs: list[str] = []  # cell addresses with external refs
        broken_refs: list[str] = []  # cell addresses with broken sheet refs
        stale: list[str] = []  # cell addresses with placeholder text

        for i, row in enumerate(ws.iter_rows(values_only=False)):
            sample_row: list[str] = []
            for cell in row:
                if cell.value is None:
                    sample_row.append("")
                    continue

                populated += 1
                value = cell.value
                cell_addr = f"{ws.title}!{cell.coordinate}"

                if isinstance(value, str) and value.startswith("="):
                    formula_cells += 1
                    formula = value

                    # --- External reference detection -----------------------
                    ext_file = _external_file_from_formula(formula)
                    if ext_file:
                        if len(ext_refs) < _LIST_CAP:
                            ext_refs.append(cell.coordinate)
                        risks.append(
                            WorkbookRisk(
                                label="External workbook reference",
                                severity="medium",
                                location=cell_addr,
                                summary=f"Formula references external file '{ext_file}'.",
                            )
                        )

                    # --- Broken reference detection -------------------------
                    tokens = _extract_range_tokens(formula)
                    if not tokens:
                        # Fallback: regex extraction of sheet names
                        tokens = [m.group(0) for m in _SHEET_REF_RE.finditer(formula)]
                    ref_sheets = _sheet_names_from_tokens(tokens)
                    for ref_sheet in ref_sheets:
                        if ref_sheet not in known_sheets:
                            if len(broken_refs) < _LIST_CAP:
                                broken_refs.append(cell.coordinate)
                            risks.append(
                                WorkbookRisk(
                                    label="Broken sheet reference",
                                    severity="high",
                                    location=cell_addr,
                                    summary=f"Formula references sheet '{ref_sheet}' which does not exist.",
                                )
                            )
                            break  # one risk per cell is enough

                    # --- Dependency graph -----------------------------------
                    if len(dependencies) < _DEP_CAP:
                        dep_cells: list[str] = []
                        for tok in _extract_range_tokens(formula):
                            # Skip external-workbook tokens
                            if "[" in tok:
                                continue
                            if "!" in tok:
                                # Already fully qualified: Sheet!CellOrRange
                                dep_cells.append(tok)
                            else:
                                # Local ref — qualify with current sheet
                                dep_cells.append(f"{ws.title}!{tok}")
                        if dep_cells:
                            dependencies[cell_addr] = dep_cells

                elif isinstance(value, str):
                    # --- Stale-input marker detection -----------------------
                    stripped = value.strip().lower()
                    if stripped in _STALE_VALUES:
                        if len(stale) < _LIST_CAP:
                            stale.append(cell.coordinate)
                        risks.append(
                            WorkbookRisk(
                                label="Pending input",
                                severity="low",
                                location=cell_addr,
                                summary=f"Cell contains placeholder text '{value.strip()}'.",
                            )
                        )

                sample_row.append(str(value))

            if i < 5:
                sample.append(sample_row[:8])

        sheets.append(
            WorkbookSheet(
                name=ws.title,
                rows=rows,
                columns=cols,
                formula_cells=formula_cells,
                populated_cells=populated,
                sample_rows=sample,
                external_references=ext_refs,
                broken_references=broken_refs,
                stale_markers=stale,
            )
        )

    # --- Named ranges -------------------------------------------------------
    named_ranges: list[WorkbookNamedRange] = []
    try:
        for nr_name, nr in book.defined_names.items():
            try:
                dests = list(nr.destinations)
            except Exception:
                dests = []
            if dests:
                sheet_name, ref = dests[0]
                full_ref = f"{sheet_name}!{ref}" if sheet_name else ref
                named_ranges.append(
                    WorkbookNamedRange(
                        name=nr_name,
                        sheet_name=sheet_name or None,
                        reference=full_ref,
                    )
                )
            else:
                named_ranges.append(
                    WorkbookNamedRange(
                        name=nr_name,
                        sheet_name=None,
                        reference=getattr(nr, "attr_text", ""),
                    )
                )
    except Exception:
        pass  # defined_names not available or malformed — skip gracefully

    wb.sheets = sheets
    wb.risks = risks
    wb.named_ranges = named_ranges
    wb.dependencies = dependencies
    return wb
