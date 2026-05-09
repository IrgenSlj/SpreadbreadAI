"""Parser tests: real risk detection, named ranges, dependency graph."""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook as XlsxWorkbook
from openpyxl.workbook.defined_name import DefinedName

from spreadbread_core.parser import parse_xlsx


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


def _save(wb: XlsxWorkbook, tmp_path: Path, name: str = "test.xlsx") -> Path:
    out = tmp_path / name
    wb.save(out)
    return out


# ---------------------------------------------------------------------------
# External reference detection
# ---------------------------------------------------------------------------


def test_parser_detects_external_reference(tmp_path: Path) -> None:
    wb = XlsxWorkbook()
    ws = wb.active
    ws.title = "Forecast"
    ws["A1"] = "=[Other.xlsx]Sheet1!$A$1"
    ws["B1"] = 100

    path = _save(wb, tmp_path)
    result = parse_xlsx(path)

    # One risk labelled "External workbook reference"
    ext_risks = [r for r in result.risks if r.label == "External workbook reference"]
    assert len(ext_risks) == 1, f"expected 1 external-ref risk, got {len(ext_risks)}"
    assert ext_risks[0].severity == "medium"
    assert "Other.xlsx" in ext_risks[0].summary

    # Sheet-level tracking
    forecast_sheet = next(s for s in result.sheets if s.name == "Forecast")
    assert len(forecast_sheet.external_references) == 1
    assert "A1" in forecast_sheet.external_references[0]


# ---------------------------------------------------------------------------
# Broken reference detection
# ---------------------------------------------------------------------------


def test_parser_detects_broken_sheet_reference(tmp_path: Path) -> None:
    wb = XlsxWorkbook()
    ws = wb.active
    ws.title = "Forecast"
    # References sheet "Pipeline" which does not exist in this workbook
    ws["A1"] = "=Pipeline!A1"
    ws["B1"] = 200

    path = _save(wb, tmp_path)
    result = parse_xlsx(path)

    broken_risks = [r for r in result.risks if r.label == "Broken sheet reference"]
    assert len(broken_risks) == 1, f"expected 1 broken-ref risk, got {len(broken_risks)}"
    assert broken_risks[0].severity == "high"
    assert "Pipeline" in broken_risks[0].summary
    assert "does not exist" in broken_risks[0].summary

    forecast_sheet = next(s for s in result.sheets if s.name == "Forecast")
    assert len(forecast_sheet.broken_references) == 1


# ---------------------------------------------------------------------------
# Stale-input marker detection
# ---------------------------------------------------------------------------


@pytest.mark.parametrize("marker", ["TODO", "TBD", "FIXME", "?", "???", "XXX", "todo", "tbd"])
def test_parser_detects_stale_marker(tmp_path: Path, marker: str) -> None:
    wb = XlsxWorkbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = marker
    ws["B1"] = 100

    path = _save(wb, tmp_path, name=f"test_{marker.lower()}.xlsx")
    result = parse_xlsx(path)

    pending_risks = [r for r in result.risks if r.label == "Pending input"]
    assert len(pending_risks) == 1, f"expected 1 pending-input risk for marker={marker!r}"
    assert pending_risks[0].severity == "low"
    assert marker.strip() in pending_risks[0].summary

    sheet = next(s for s in result.sheets if s.name == "Sheet1")
    assert len(sheet.stale_markers) == 1


# ---------------------------------------------------------------------------
# Named range extraction
# ---------------------------------------------------------------------------


def test_parser_extracts_named_ranges(tmp_path: Path) -> None:
    wb = XlsxWorkbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = 100
    ws["B10"] = 200

    dn = DefinedName("RevenueRange", attr_text="Sheet1!$A$1:$B$10")
    wb.defined_names.add(dn)

    path = _save(wb, tmp_path)
    result = parse_xlsx(path)

    assert len(result.named_ranges) >= 1
    names = {nr.name for nr in result.named_ranges}
    assert "RevenueRange" in names

    rev_range = next(nr for nr in result.named_ranges if nr.name == "RevenueRange")
    assert "Sheet1" in rev_range.reference
    assert "$A$1" in rev_range.reference or "A1" in rev_range.reference


# ---------------------------------------------------------------------------
# Dependency graph
# ---------------------------------------------------------------------------


def test_parser_builds_dependency_graph(tmp_path: Path) -> None:
    wb = XlsxWorkbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = 10
    ws["B2"] = 20
    ws["C3"] = "=A1+B2"

    path = _save(wb, tmp_path)
    result = parse_xlsx(path)

    key = "Sheet1!C3"
    assert key in result.dependencies, f"{key} not found in dependencies: {list(result.dependencies.keys())}"
    deps = result.dependencies[key]
    assert "Sheet1!A1" in deps
    assert "Sheet1!B2" in deps


# ---------------------------------------------------------------------------
# Placeholder risk is removed
# ---------------------------------------------------------------------------


def test_parser_no_more_placeholder_risks(tmp_path: Path) -> None:
    """A workbook with formulas but no external refs, broken refs, or stale
    markers must produce zero risks — the old placeholder is gone."""
    wb = XlsxWorkbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = 100
    ws["B1"] = 200
    ws["C1"] = "=A1+B1"  # valid intra-sheet formula, no issues

    path = _save(wb, tmp_path)
    result = parse_xlsx(path)

    placeholder_risks = [r for r in result.risks if "Formula review pending" in r.label]
    assert placeholder_risks == [], f"placeholder risk still present: {placeholder_risks}"
    assert result.risks == [], f"expected no risks, got: {result.risks}"


# ---------------------------------------------------------------------------
# Dependency graph cap
# ---------------------------------------------------------------------------


def test_parser_dependency_graph_caps_at_500(tmp_path: Path) -> None:
    wb = XlsxWorkbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = 1
    # Fill 600 cells with =A1
    for row in range(2, 602):
        ws.cell(row=row, column=2).value = "=A1"

    path = _save(wb, tmp_path)
    result = parse_xlsx(path)

    assert len(result.dependencies) <= 500, (
        f"dependency graph exceeds 500 entries: {len(result.dependencies)}"
    )
