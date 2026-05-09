"""End-to-end smoke test: parse a tiny xlsx, run the registry, verify
proposal items are staged (not auto-applied) and audit events accrete.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook as XlsxWorkbook

from spreadbread_core.config import Config
from spreadbread_core.parser import parse_xlsx
from spreadbread_core.store import Store
from spreadbread_core.tools import ToolRegistry


def _build_sample_xlsx(tmp_path: Path) -> Path:
    book = XlsxWorkbook()
    sheet = book.active
    sheet.title = "Forecast"
    sheet.append(["Month", "Quota", "Forecast"])
    sheet.append(["Apr", 500000, 478000])
    sheet.append(["May", 520000, "=B3*1.05"])
    out = tmp_path / "sample.xlsx"
    book.save(out)
    return out


def test_end_to_end(tmp_path: Path) -> None:
    db_path = tmp_path / "test.sqlite3"
    store = Store(db_path)
    registry = ToolRegistry(store)

    # parse + save
    xlsx_path = _build_sample_xlsx(tmp_path)
    wb = parse_xlsx(xlsx_path)
    store.save_workbook(wb)
    assert wb.sheets[0].name == "Forecast"
    assert wb.sheets[0].formula_cells == 1

    # tool: list_workbooks
    listed = registry.call("list_workbooks", {})
    assert len(listed) == 1
    assert listed[0]["id"] == wb.id

    # tool: inspect_sheet
    sheet = registry.call("inspect_sheet", {"workbook_id": wb.id, "sheet_name": "Forecast"})
    assert sheet["formula_cells"] == 1

    # tool: propose_diff stages an item, does NOT mutate workbook
    result = registry.call(
        "propose_diff",
        {
            "workbook_id": wb.id,
            "cell": "Forecast!C3",
            "kind": "update",
            "before": "=B3*1.05",
            "after": "=B3*1.08",
            "rationale": "Aligns May forecast with revised growth assumption.",
        },
    )
    assert result["status"] == "pending"

    snap = store.review_snapshot(wb.id)
    assert snap is not None
    assert snap.proposal is not None
    assert len(snap.proposal.items) == 1
    item = snap.proposal.items[0]
    assert item.status == "pending"  # human approval still required
    assert item.cell == "Forecast!C3"

    # decision: approve
    proposal = store.decide_item(snap.proposal.id, item.id, "approve", reviewer="finance_manager")
    assert proposal.items[0].status == "approved"

    # audit accreted: upload? no — parser only writes via http path.
    # but tool flow appended proposal.created + proposal.item.added
    audit = store.list_audit(wb.id)
    actions = {e.action for e in audit}
    assert "proposal.created" in actions
    assert "proposal.item.added" in actions


def _build_enriched_xlsx(tmp_path: Path) -> Path:
    """Workbook with a named range, an external reference, and a formula."""
    from openpyxl import Workbook as XlsxWorkbook
    from openpyxl.workbook.defined_name import DefinedName

    book = XlsxWorkbook()
    ws = book.active
    ws.title = "Forecast"
    ws["A1"] = 100
    ws["B1"] = 200
    ws["C1"] = "=A1+B1"
    ws["D1"] = "=[Other.xlsx]Sheet1!$A$1"

    dn = DefinedName("MyRange", attr_text="Forecast!$A$1:$B$1")
    book.defined_names.add(dn)

    out = tmp_path / "enriched.xlsx"
    book.save(out)
    return out


def test_new_tools_end_to_end(tmp_path: Path) -> None:
    """get_dependencies, find_external_references, get_named_ranges return expected shapes."""
    db_path = tmp_path / "tools_test.sqlite3"
    store = Store(db_path)
    registry = ToolRegistry(store)

    xlsx_path = _build_enriched_xlsx(tmp_path)
    wb = parse_xlsx(xlsx_path)
    store.save_workbook(wb)

    # get_dependencies — C1 depends on A1 and B1
    dep_result = registry.call("get_dependencies", {"workbook_id": wb.id, "cell": "Forecast!C1"})
    assert dep_result["cell"] == "Forecast!C1"
    assert isinstance(dep_result["depends_on"], list)
    assert "Forecast!A1" in dep_result["depends_on"]
    assert "Forecast!B1" in dep_result["depends_on"]

    # get_dependencies — unknown cell returns empty list
    dep_none = registry.call("get_dependencies", {"workbook_id": wb.id, "cell": "Forecast!Z99"})
    assert dep_none["depends_on"] == []

    # find_external_references — D1 has an external ref
    ext_result = registry.call("find_external_references", {"workbook_id": wb.id})
    assert isinstance(ext_result, list)
    assert len(ext_result) >= 1
    sheets_returned = {e["sheet"] for e in ext_result}
    assert "Forecast" in sheets_returned

    # get_named_ranges — MyRange should be present
    nr_result = registry.call("get_named_ranges", {"workbook_id": wb.id})
    assert isinstance(nr_result, list)
    assert len(nr_result) >= 1
    names = {nr["name"] for nr in nr_result}
    assert "MyRange" in names
    my_range = next(nr for nr in nr_result if nr["name"] == "MyRange")
    assert "reference" in my_range
    assert "name" in my_range
    assert "sheet_name" in my_range


def test_config_loads(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setenv("SPREADBREAD_DATA_DIR", str(tmp_path))
    cfg = Config.load()
    assert cfg.db_path.parent == tmp_path
    assert cfg.model  # default present
