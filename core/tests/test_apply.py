"""Tests for the apply pipeline."""
from __future__ import annotations

import io
from pathlib import Path

import pytest
from openpyxl import Workbook as XlsxWorkbook
from openpyxl import load_workbook

from spreadbread_core.apply import ApplyError, apply_proposal
from spreadbread_core.domain import ProposalItem, new_proposal
from spreadbread_core.parser import parse_xlsx
from spreadbread_core.store import Store


def _seed(tmp_path: Path) -> tuple[Store, str, str]:
    """Create store + sample workbook with bytes; return (store, workbook_id, proposal_id)."""
    store = Store(tmp_path / "apply.sqlite3")

    book = XlsxWorkbook()
    sheet = book.active
    sheet.title = "Forecast"
    sheet.append(["Month", "Quota", "Forecast"])
    sheet.append(["Apr", 500000, 478000])
    sheet.append(["May", 520000, "=B3*1.05"])
    out = tmp_path / "sample.xlsx"
    book.save(out)
    raw = out.read_bytes()

    wb = parse_xlsx(out)
    store.save_workbook(wb)
    store.save_version_bytes(wb.id, wb.latest_version_id, raw)

    proposal = new_proposal(
        wb.id, "test", "test", "llm",
        source_version_id=wb.latest_version_id,
        source_version_sha256=store.version_sha256(wb.id, wb.latest_version_id),
    )
    proposal.items = [
        ProposalItem(kind="update", cell="Forecast!C3", before="=B3*1.05",
                     after="=B3*1.08", rationale="growth", status="approved"),
        ProposalItem(kind="comment", cell="Forecast!A1", after="Reviewed by test",
                     rationale="note", status="approved"),
        ProposalItem(kind="update", cell="Forecast!B2", before="500000",
                     after="510000", rationale="quota", status="rejected"),
    ]
    store.save_proposal(proposal)
    return store, wb.id, proposal.id


def test_apply_writes_new_version(tmp_path: Path) -> None:
    store, wb_id, prop_id = _seed(tmp_path)

    result = apply_proposal(store, prop_id, reviewer="finance")

    # proposal flipped to applied
    assert result.proposal.status == "applied"
    assert result.proposal.applied_version_id == result.version.id
    assert len(result.applied_item_ids) == 2  # rejected item not applied

    # workbook has new version, latest_version_id updated
    workbook = store.get_workbook(wb_id)
    assert workbook is not None
    assert workbook.latest_version_id == result.version.id
    assert any(v.id == result.version.id for v in workbook.versions)
    assert workbook.status == "healthy"

    # bytes were written and reflect approved items
    raw_new = store.load_version_bytes(wb_id, result.version.id)
    book = load_workbook(io.BytesIO(raw_new), data_only=False)
    sheet = book["Forecast"]
    assert sheet["C3"].value == "=B3*1.08"
    assert sheet["B2"].value == 500000  # rejected item was NOT applied
    assert sheet["A1"].comment is not None
    assert "Reviewed by test" in sheet["A1"].comment.text

    # audit accreted
    actions = {e.action for e in store.list_audit(wb_id)}
    assert "proposal.applied" in actions
    assert "version.created" in actions


def test_apply_is_idempotent(tmp_path: Path) -> None:
    store, _, prop_id = _seed(tmp_path)
    first = apply_proposal(store, prop_id, reviewer="finance")
    second = apply_proposal(store, prop_id, reviewer="finance")
    assert first.version.id == second.version.id


def test_apply_rejects_pending_items(tmp_path: Path) -> None:
    store, _, prop_id = _seed(tmp_path)
    proposal = store.get_proposal(prop_id)
    assert proposal is not None
    proposal.items[0].status = "pending"
    store.save_proposal(proposal)
    with pytest.raises(ApplyError, match="pending"):
        apply_proposal(store, prop_id)


def test_apply_rejects_when_no_approved_items(tmp_path: Path) -> None:
    store, _, prop_id = _seed(tmp_path)
    proposal = store.get_proposal(prop_id)
    assert proposal is not None
    for item in proposal.items:
        item.status = "rejected"
    store.save_proposal(proposal)
    with pytest.raises(ApplyError, match="no approved"):
        apply_proposal(store, prop_id)


def test_apply_refuses_when_workbook_moved(tmp_path: Path) -> None:
    store, wb_id, prop_id = _seed(tmp_path)
    # Simulate the workbook having advanced to a new version after the
    # proposal was created (e.g. user re-uploaded).
    workbook = store.get_workbook(wb_id)
    assert workbook is not None
    workbook.latest_version_id = "wbv_someone_else"
    store.save_workbook(workbook)
    with pytest.raises(ApplyError, match="moved"):
        apply_proposal(store, prop_id)


def test_apply_refuses_when_base_bytes_tampered(tmp_path: Path) -> None:
    store, wb_id, prop_id = _seed(tmp_path)
    workbook = store.get_workbook(wb_id)
    assert workbook is not None
    # Same version id but different bytes.
    store.save_version_bytes(wb_id, workbook.latest_version_id, b"tampered")
    with pytest.raises(ApplyError, match="sha256"):
        apply_proposal(store, prop_id)


def test_decide_all_pending_flips_only_pending(tmp_path: Path) -> None:
    store, _, prop_id = _seed(tmp_path)
    proposal = store.get_proposal(prop_id)
    assert proposal is not None
    # _seed leaves: 2 approved, 1 rejected, 0 pending
    proposal.items[0].status = "pending"
    proposal.items[1].status = "pending"
    store.save_proposal(proposal)

    updated, flipped = store.decide_all_pending(prop_id, "approve", reviewer="finance")
    assert len(flipped) == 2
    statuses = [item.status for item in updated.items]
    assert statuses.count("approved") == 2
    assert statuses.count("rejected") == 1
    for item in updated.items:
        if item.id in flipped:
            assert item.reviewer == "finance"
            assert item.reviewed_at is not None
