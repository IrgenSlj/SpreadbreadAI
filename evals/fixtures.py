"""Synthetic workbook fixtures for the eval harness.

Each function returns a tuple of ``(name, bytes)`` suitable for uploading
through the store/parser pipeline. No external files needed.
"""
from __future__ import annotations

import io
from typing import Callable

from openpyxl import Workbook as XlsxWorkbook

FixtureFn = Callable[[], tuple[str, bytes]]
_registry: dict[str, FixtureFn] = {}


def _register(fn: FixtureFn) -> FixtureFn:
    _registry[fn.__name__] = fn
    return fn


def list_fixtures() -> list[str]:
    return list(_registry.keys())


def load(name: str) -> tuple[str, bytes]:
    if name not in _registry:
        raise KeyError(f"unknown fixture: {name!r} (available: {list(_registry)})")
    return _registry[name]()


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@_register
def simple_forecast() -> tuple[str, bytes]:
    """A 3-column, 4-row forecast workbook: Month | Quota | Forecast."""
    book = XlsxWorkbook()
    ws = book.active
    ws.title = "Forecast"
    ws.append(["Month", "Quota", "Forecast"])
    ws.append(["Apr", 500000, 478000])
    ws.append(["May", 520000, "=B3*1.05"])
    ws.append(["Jun", 550000, "=B4*1.03"])
    buf = io.BytesIO()
    book.save(buf)
    return "simple_forecast", buf.getvalue()


@_register
def construction_quote() -> tuple[str, bytes]:
    """A construction cost-estimate workbook with multiple tabs.

    Sheets: Summary, Labor, Materials, Equipment, Quote.
    Mix of values and formulas across sheets.
    """
    book = XlsxWorkbook()

    # --- Summary ---
    ws = book.active
    ws.title = "Summary"
    ws.append(["Item", "Cost", "Margin"])
    ws.append(["Labor", 45000, "=B2*0.15"])
    ws.append(["Materials", 32000, "=B3*0.12"])
    ws.append(["Equipment", 18000, "=B4*0.10"])
    ws.append(["Total", "=SUM(B2:B4)", "=AVERAGE(C2:C4)"])

    # --- Labor ---
    ws2 = book.create_sheet("Labor")
    ws2.append(["Role", "Hours", "Rate", "Total"])
    ws2.append(["Foreman", 40, 65, "=B2*C2"])
    ws2.append(["Carpenter", 80, 55, "=B3*C3"])
    ws2.append(["Laborer", 120, 35, "=B4*C4"])

    # --- Materials ---
    ws3 = book.create_sheet("Materials")
    ws3.append(["Item", "Qty", "Unit Price", "Total"])
    ws3.append(["Lumber", 200, 45, "=B2*C2"])
    ws3.append(["Concrete", 15, 120, "=B3*C3"])
    ws3.append(["Roofing", 100, 65, "=B4*C4"])

    # --- Equipment ---
    ws4 = book.create_sheet("Equipment")
    ws4.append(["Equipment", "Days", "Daily Rate", "Total"])
    ws4.append(["Excavator", 5, 850, "=B2*C2"])
    ws4.append(["Forklift", 3, 400, "=B3*C3"])

    # --- Quote (with external ref to Summary) ---
    ws5 = book.create_sheet("Quote")
    ws5.append(["Line", "Amount"])
    ws5.append(["Labor Total", "=Summary!B2"])
    ws5.append(["Materials Total", "=Summary!B3"])
    ws5.append(["Equipment Total", "=Summary!B4"])
    ws5.append(["Subtotal", "=SUM(B2:B4)"])
    ws5.append(["Contingency", "=B5*0.1"])
    ws5.append(["Total", "=B5+B6"])

    buf = io.BytesIO()
    book.save(buf)
    return "construction_quote", buf.getvalue()


@_register
def broken_refs() -> tuple[str, bytes]:
    """A workbook with known issues: external refs, stale markers, broken sheet refs."""
    book = XlsxWorkbook()
    ws = book.active
    ws.title = "Sheet1"
    ws["A1"] = "Item"
    ws["A2"] = "OK"
    ws["B1"] = 100
    ws["C1"] = "=B1+50"
    ws["D1"] = "=[External.xlsx]Sheet1!A1"  # external reference
    ws["E1"] = "=Missing!A1"  # broken sheet ref
    ws["F1"] = "TBD"  # stale marker
    ws["G1"] = "XXX"  # stale marker
    buf = io.BytesIO()
    book.save(buf)
    return "broken_refs", buf.getvalue()


@_register
def cycle_risk() -> tuple[str, bytes]:
    """A workbook with a formula that contains a broken cross-sheet reference
    (which the parser's risk detection should catch)."""
    book = XlsxWorkbook()
    ws = book.active
    ws.title = "Sheet1"
    ws["A1"] = 10
    ws["A2"] = 20
    ws["A3"] = "=A1+A2"
    ws["B1"] = "=MissingSheet!A1+1"  # broken sheet ref (detectable by parser)
    buf = io.BytesIO()
    book.save(buf)
    return "cycle_risk", buf.getvalue()
